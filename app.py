from flask import Flask, render_template, request, redirect, url_for, session, send_file, jsonify
import sqlite3, json, os, io, base64
from datetime import datetime
from functools import wraps

app = Flask(__name__)
app.secret_key = "okul_anket_super_gizli_2024"
DB = "anket.db"

def db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    return con

def db_init():
    with db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS ayarlar (
            anahtar TEXT PRIMARY KEY, deger TEXT
        );
        CREATE TABLE IF NOT EXISTS anketler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rol TEXT NOT NULL, baslik TEXT NOT NULL,
            icon TEXT NOT NULL DEFAULT '📝',
            aktif INTEGER DEFAULT 1, sira INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS bolumler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            anket_id INTEGER NOT NULL, baslik TEXT NOT NULL, sira INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS sorular (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bolum_id INTEGER NOT NULL, metin TEXT NOT NULL,
            tip TEXT NOT NULL DEFAULT 'yildiz',
            secenekler TEXT DEFAULT '[]',
            zorunlu INTEGER DEFAULT 1, sira INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS yanitlar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            anket_id INTEGER NOT NULL,
            tarih TEXT NOT NULL, saat TEXT NOT NULL, veriler TEXT NOT NULL
        );
        """)
        defaults = {
            "okul_adi": "Örnek İlköğretim Okulu",
            "okul_sehir": "Erzurum",
            "admin_sifre": "okul2024",
            "tema": "mavi",
            "amblem": "",
            "anasayfa_gorunum": "kartlar",
            "hosgeldin_metin": "Görüşleriniz okulumuzun gelişimine katkı sağlar.",
            "alt_yazi": "Anketler anonim olarak kaydedilir."
        }
        for k, v in defaults.items():
            c.execute("INSERT OR IGNORE INTO ayarlar VALUES (?,?)", (k, v))
        if not c.execute("SELECT 1 FROM anketler").fetchone():
            _varsayilan_anketler(c)
        c.commit()

def _varsayilan_anketler(c):
    veri = [
        ("ogrenci","Öğrenci Değerlendirme Anketi","🎒",0,[
            ("Ders ve Öğretim",[
                ("Derslerin işlenişini nasıl değerlendirirsiniz?","yildiz","[]"),
                ("Öğretmenler sorularınıza yeterince ilgi gösteriyor mu?","secim",'["Her zaman","Çoğunlukla","Bazen","Nadiren"]'),
            ]),
            ("Okul Ortamı",[
                ("Okul temizliği ve düzenini nasıl buluyorsunuz?","yildiz","[]"),
                ("Okul ortamında kendinizi güvende hissediyor musunuz?","secim",'["Evet, her zaman","Genellikle evet","Bazen hayır","Hayır"]'),
            ]),
            ("Genel Değerlendirme",[
                ("Okulunuzdan genel memnuniyetiniz nedir?","yildiz","[]"),
                ("Okulun daha iyi olması için öneriniz var mı?","metin","[]"),
            ]),
        ]),
        ("veli","Veli Değerlendirme Anketi","👨‍👩‍👧",1,[
            ("İletişim ve Bilgilendirme",[
                ("Okul ile iletişim kalitesini nasıl değerlendirirsiniz?","yildiz","[]"),
                ("Çocuğunuzun gelişimi hakkında yeterince bilgilendiriliyor musunuz?","secim",'["Evet, düzenli olarak","Kısmen","Hayır, yeterince değil"]'),
            ]),
            ("Eğitim Kalitesi",[
                ("Okulun eğitim kalitesini nasıl değerlendirirsiniz?","yildiz","[]"),
                ("Öğretmenler velilere karşı destekleyici mi?","secim",'["Evet","Çoğunlukla","Bazen","Hayır"]'),
            ]),
            ("Genel Görüş",[
                ("Okuldan genel memnuniyetiniz?","yildiz","[]"),
                ("Paylaşmak istediğiniz görüş veya öneri:","metin","[]"),
            ]),
        ]),
        ("ogretmen","Öğretmen Değerlendirme Anketi","📚",2,[
            ("Yönetim ve Destek",[
                ("Okul yönetiminin öğretmenlere desteğini nasıl değerlendirirsiniz?","yildiz","[]"),
                ("Okul kararlarında görüşleriniz dikkate alınıyor mu?","secim",'["Her zaman","Çoğunlukla","Bazen","Nadiren"]'),
            ]),
            ("Çalışma Ortamı",[
                ("Fiziksel çalışma ortamını nasıl değerlendirirsiniz?","yildiz","[]"),
                ("Ders materyalleri ve kaynaklar yeterli mi?","secim",'["Evet, çok yeterli","Kısmen yeterli","Yetersiz"]'),
            ]),
            ("Genel Değerlendirme",[
                ("İş memnuniyetiniz genel olarak nasıl?","yildiz","[]"),
                ("Okulu geliştirmek için önerileriniz:","metin","[]"),
            ]),
        ]),
    ]
    for rol,baslik,icon,sira,bolumler in veri:
        aid=c.execute("INSERT INTO anketler (rol,baslik,icon,sira) VALUES (?,?,?,?)",(rol,baslik,icon,sira)).lastrowid
        for bs,(bb,sorular) in enumerate(bolumler):
            bid=c.execute("INSERT INTO bolumler (anket_id,baslik,sira) VALUES (?,?,?)",(aid,bb,bs)).lastrowid
            for ss,(sm,st,sse) in enumerate(sorular):
                c.execute("INSERT INTO sorular (bolum_id,metin,tip,secenekler,sira) VALUES (?,?,?,?,?)",(bid,sm,st,sse,ss))

db_init()

def ayar(k, varsayilan=""):
    with db() as c:
        r=c.execute("SELECT deger FROM ayarlar WHERE anahtar=?",(k,)).fetchone()
    return r["deger"] if r else varsayilan

def ayar_set(k,v):
    with db() as c:
        c.execute("INSERT OR REPLACE INTO ayarlar VALUES (?,?)",(k,v))
        c.commit()

def get_anket(anket_id):
    with db() as c:
        a=c.execute("SELECT * FROM anketler WHERE id=?",(anket_id,)).fetchone()
        if not a: return None
        bolumler=c.execute("SELECT * FROM bolumler WHERE anket_id=? ORDER BY sira",(anket_id,)).fetchall()
        result=dict(a); result["bolumler"]=[]
        for b in bolumler:
            bd=dict(b)
            sorular=c.execute("SELECT * FROM sorular WHERE bolum_id=? ORDER BY sira",(b["id"],)).fetchall()
            bd["sorular"]=[dict(s) for s in sorular]
            for s in bd["sorular"]:
                s["secenekler"]=json.loads(s["secenekler"] or "[]")
            result["bolumler"].append(bd)
        return result

def gctx():
    return {"okul_adi":ayar("okul_adi"),"okul_sehir":ayar("okul_sehir"),"tema":ayar("tema","mavi"),"amblem":ayar("amblem")}

def giris_gerekli(f):
    @wraps(f)
    def k(*a,**kw):
        if not session.get("admin"): return redirect(url_for("admin_giris"))
        return f(*a,**kw)
    return k

# ── Anasayfa & Anket ─────────────────────────────────────────────
@app.route("/")
def anasayfa():
    with db() as c:
        anketler=c.execute("SELECT * FROM anketler WHERE aktif=1 ORDER BY sira").fetchall()
    ctx=gctx(); ctx["anketler"]=[dict(a) for a in anketler]
    ctx["hosgeldin"]=ayar("hosgeldin_metin"); ctx["alt_yazi"]=ayar("alt_yazi")
    ctx["gorunum"]=ayar("anasayfa_gorunum","kartlar")
    return render_template("anasayfa.html",**ctx)

@app.route("/anket/<int:anket_id>",methods=["GET","POST"])
def anket(anket_id):
    a=get_anket(anket_id)
    if not a or not a["aktif"]: return redirect(url_for("anasayfa"))
    if request.method=="POST":
        v={k:val for k,val in request.form.items()}
        with db() as c:
            c.execute("INSERT INTO yanitlar (anket_id,tarih,saat,veriler) VALUES (?,?,?,?)",
                      (anket_id,datetime.now().strftime("%d.%m.%Y"),datetime.now().strftime("%H:%M"),
                       json.dumps(v,ensure_ascii=False)))
            c.commit()
        return redirect(url_for("tesekkur",anket_id=anket_id))
    ctx=gctx(); ctx["anket"]=a
    return render_template("anket.html",**ctx)

@app.route("/tesekkur/<int:anket_id>")
def tesekkur(anket_id):
    with db() as c:
        a=c.execute("SELECT * FROM anketler WHERE id=?",(anket_id,)).fetchone()
    ctx=gctx(); ctx["anket"]=dict(a) if a else {"baslik":"Anket","icon":"✅","id":0}
    return render_template("tesekkur.html",**ctx)

# ── Admin Giriş ──────────────────────────────────────────────────
@app.route("/admin",methods=["GET","POST"])
def admin_giris():
    hata=None
    if request.method=="POST":
        if request.form.get("sifre")==ayar("admin_sifre","okul2024"):
            session["admin"]=True; return redirect(url_for("admin_panel"))
        hata="Yanlış şifre!"
    return render_template("admin_giris.html",hata=hata,**gctx())

@app.route("/admin/cikis")
def admin_cikis():
    session.clear(); return redirect(url_for("anasayfa"))

# ── Admin Panel ───────────────────────────────────────────────────
@app.route("/admin/panel")
@giris_gerekli
def admin_panel():
    with db() as c:
        anketler=c.execute("SELECT * FROM anketler ORDER BY sira").fetchall()
        istat=[]
        toplam=0
        for a in anketler:
            sayi=c.execute("SELECT COUNT(*) as n FROM yanitlar WHERE anket_id=?",(a["id"],)).fetchone()["n"]
            toplam+=sayi
            yanitlar=c.execute("SELECT veriler FROM yanitlar WHERE anket_id=?",(a["id"],)).fetchall()
            yy={}
            for y in yanitlar:
                for k,val in json.loads(y["veriler"]).items():
                    if k.startswith("s_"):
                        try:
                            iv=int(val)
                            if 1<=iv<=5: yy.setdefault(k,[]).append(iv)
                        except: pass
            avg={k:round(sum(vs)/len(vs),1) for k,vs in yy.items()}
            istat.append({"anket":dict(a),"sayi":sayi,"avg":avg})
        son=c.execute("""SELECT y.*,a.baslik as ab,a.icon as ai FROM yanitlar y
            JOIN anketler a ON y.anket_id=a.id ORDER BY y.id DESC LIMIT 15""").fetchall()
    ctx=gctx(); ctx.update({"istat":istat,"toplam":toplam,"son":[dict(r) for r in son]})
    return render_template("admin_panel.html",**ctx)

# ── Sonuçlar ─────────────────────────────────────────────────────
@app.route("/admin/sonuclar/<int:anket_id>")
@giris_gerekli
def admin_sonuclar(anket_id):
    a=get_anket(anket_id)
    if not a: return redirect(url_for("admin_panel"))
    with db() as c:
        yanitlar=c.execute("SELECT * FROM yanitlar WHERE anket_id=? ORDER BY id DESC",(anket_id,)).fetchall()
    liste=[{"id":y["id"],"tarih":y["tarih"],"saat":y["saat"],"veriler":json.loads(y["veriler"])} for y in yanitlar]
    ctx=gctx(); ctx.update({"anket":a,"yanitlar":liste})
    return render_template("admin_sonuclar.html",**ctx)

@app.route("/admin/yanit_sil/<int:yid>",methods=["POST"])
@giris_gerekli
def yanit_sil(yid):
    with db() as c:
        c.execute("DELETE FROM yanitlar WHERE id=?",(yid,)); c.commit()
    return redirect(request.referrer or url_for("admin_panel"))

# ── Anket CRUD ────────────────────────────────────────────────────
@app.route("/admin/anketler")
@giris_gerekli
def admin_anketler():
    with db() as c:
        anketler=c.execute("SELECT * FROM anketler ORDER BY sira").fetchall()
    ctx=gctx(); ctx["anketler"]=[dict(a) for a in anketler]
    return render_template("admin_anketler.html",**ctx)

@app.route("/admin/anket/yeni",methods=["POST"])
@giris_gerekli
def anket_yeni():
    with db() as c:
        aid=c.execute("INSERT INTO anketler (rol,baslik,icon,aktif,sira) VALUES (?,?,?,1,99)",
                      (request.form.get("rol","diger"),request.form.get("baslik","Yeni Anket"),request.form.get("icon","📝"))).lastrowid
        c.execute("INSERT INTO bolumler (anket_id,baslik,sira) VALUES (?,?,0)",(aid,"Genel Sorular"))
        c.commit()
    return redirect(url_for("admin_anket_duzenle",anket_id=aid))

@app.route("/admin/anket/<int:aid>/duzenle")
@giris_gerekli
def admin_anket_duzenle(aid):
    a=get_anket(aid)
    if not a: return redirect(url_for("admin_anketler"))
    ctx=gctx(); ctx["anket"]=a
    return render_template("admin_anket_duzenle.html",**ctx)

@app.route("/admin/anket/<int:aid>/guncelle",methods=["POST"])
@giris_gerekli
def anket_guncelle(aid):
    with db() as c:
        c.execute("UPDATE anketler SET baslik=?,icon=?,rol=?,aktif=? WHERE id=?",
                  (request.form.get("baslik"),request.form.get("icon"),request.form.get("rol"),
                   1 if request.form.get("aktif") else 0,aid)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/anket/<int:aid>/sil",methods=["POST"])
@giris_gerekli
def anket_sil(aid):
    with db() as c:
        c.execute("DELETE FROM yanitlar WHERE anket_id=?",(aid,))
        for b in c.execute("SELECT id FROM bolumler WHERE anket_id=?",(aid,)).fetchall():
            c.execute("DELETE FROM sorular WHERE bolum_id=?",(b["id"],))
        c.execute("DELETE FROM bolumler WHERE anket_id=?",(aid,))
        c.execute("DELETE FROM anketler WHERE id=?",(aid,)); c.commit()
    return redirect(url_for("admin_anketler"))

# ── Bölüm CRUD ───────────────────────────────────────────────────
@app.route("/admin/bolum/ekle",methods=["POST"])
@giris_gerekli
def bolum_ekle():
    aid=int(request.form["anket_id"])
    with db() as c:
        c.execute("INSERT INTO bolumler (anket_id,baslik,sira) VALUES (?,?,99)",
                  (aid,request.form.get("baslik","Yeni Bölüm"))); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/bolum/<int:bid>/guncelle",methods=["POST"])
@giris_gerekli
def bolum_guncelle(bid):
    aid=int(request.form["anket_id"])
    with db() as c:
        c.execute("UPDATE bolumler SET baslik=? WHERE id=?",(request.form.get("baslik"),bid)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/bolum/<int:bid>/sil",methods=["POST"])
@giris_gerekli
def bolum_sil(bid):
    aid=int(request.form["anket_id"])
    with db() as c:
        c.execute("DELETE FROM sorular WHERE bolum_id=?",(bid,))
        c.execute("DELETE FROM bolumler WHERE id=?",(bid,)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

# ── Soru CRUD ────────────────────────────────────────────────────
@app.route("/admin/soru/ekle",methods=["POST"])
@giris_gerekli
def soru_ekle():
    bid=int(request.form["bolum_id"]); aid=int(request.form["anket_id"])
    tip=request.form.get("tip","yildiz")
    secs=[s.strip() for s in request.form.getlist("secenek") if s.strip()]
    with db() as c:
        c.execute("INSERT INTO sorular (bolum_id,metin,tip,secenekler,zorunlu,sira) VALUES (?,?,?,?,?,99)",
                  (bid,request.form.get("metin","Yeni soru?"),tip,json.dumps(secs,ensure_ascii=False),
                   1 if request.form.get("zorunlu") else 0)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/soru/<int:sid>/guncelle",methods=["POST"])
@giris_gerekli
def soru_guncelle(sid):
    aid=int(request.form["anket_id"]); tip=request.form.get("tip","yildiz")
    secs=[s.strip() for s in request.form.getlist("secenek") if s.strip()]
    with db() as c:
        c.execute("UPDATE sorular SET metin=?,tip=?,secenekler=?,zorunlu=? WHERE id=?",
                  (request.form.get("metin"),tip,json.dumps(secs,ensure_ascii=False),
                   1 if request.form.get("zorunlu") else 0,sid)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

@app.route("/admin/soru/<int:sid>/sil",methods=["POST"])
@giris_gerekli
def soru_sil(sid):
    aid=int(request.form["anket_id"])
    with db() as c:
        c.execute("DELETE FROM sorular WHERE id=?",(sid,)); c.commit()
    return redirect(url_for("admin_anket_duzenle",aid=aid))

# ── Ayarlar ──────────────────────────────────────────────────────
@app.route("/admin/ayarlar",methods=["GET","POST"])
@giris_gerekli
def admin_ayarlar():
    mesaj=None
    if request.method=="POST":
        for k in ["okul_adi","okul_sehir","admin_sifre","tema","anasayfa_gorunum","hosgeldin_metin","alt_yazi"]:
            v=request.form.get(k)
            if v is not None: ayar_set(k,v)
        f=request.files.get("amblem")
        if f and f.filename:
            data=f.read(); ext=f.filename.rsplit(".",1)[-1].lower()
            mime="image/png" if ext=="png" else "image/jpeg"
            ayar_set("amblem",f"data:{mime};base64,{base64.b64encode(data).decode()}")
        elif request.form.get("amblem_sil"):
            ayar_set("amblem","")
        mesaj="Ayarlar kaydedildi! ✓"
    ctx=gctx(); ctx.update({"mesaj":mesaj,"gorunum":ayar("anasayfa_gorunum","kartlar"),
                             "hosgeldin":ayar("hosgeldin_metin"),"alt_yazi":ayar("alt_yazi"),
                             "admin_sifre":ayar("admin_sifre","okul2024")})
    return render_template("admin_ayarlar.html",**ctx)

# ── Excel ────────────────────────────────────────────────────────
@app.route("/admin/excel/<int:anket_id>")
@giris_gerekli
def excel_indir(anket_id):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except: return "openpyxl kurulu değil.",500
    a=get_anket(anket_id)
    if not a: return redirect(url_for("admin_panel"))
    with db() as c:
        yanitlar=c.execute("SELECT * FROM yanitlar WHERE anket_id=? ORDER BY id",(anket_id,)).fetchall()
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=a["baslik"][:30]
    hdrs=["No","Tarih","Saat"]
    sid_order=[]
    for b in a["bolumler"]:
        for s in b["sorular"]:
            hdrs.append(s["metin"][:50]); sid_order.append(s["id"])
    for ci,h in enumerate(hdrs,1):
        cell=ws.cell(1,ci,h); cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="185FA5"); cell.alignment=Alignment(horizontal="center",wrap_text=True)
    ws.row_dimensions[1].height=40
    for ri,y in enumerate(yanitlar,2):
        v=json.loads(y["veriler"]); ws.cell(ri,1,y["id"]); ws.cell(ri,2,y["tarih"]); ws.cell(ri,3,y["saat"])
        for ci,sid in enumerate(sid_order,4):
            ws.cell(ri,ci,v.get(f"s_{sid}",""))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=22
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fn=f"{a['baslik'][:20]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fn,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    app.run(debug=False,host="0.0.0.0",port=port)
